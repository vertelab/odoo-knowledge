import openpyxl
import base64
from io import BytesIO

from odoo import models, fields, api

class ImportESGWizard(models.TransientModel):
    _name = 'import.esg.wizard'

    file = fields.Binary(string="File", required=True)

    field_keys = [
    "csrd_id", 
    "csrd_esrs", 
    "csrd_dr", 
    "csrd_paragraph",
    "csrd_related_ar",
    "csrd_name",
    "csrd_data_type",
    "csrd_conditional_or_alternative_dp",
    "csrd_may_v",
    "csrd_appendix_b",
    "csrd_appendix_c_less_then_750", 
    "csrd_appendix_c_more_then_750",
    ]

    def import_esg(self):
        
        wb = openpyxl.load_workbook(filename=BytesIO(base64.b64decode(self.file)))

        wslist = wb.sheetnames

        wslist.pop(0)

        for ws in wslist:

            activews = wb[ws]
            record = {}

            for row in range(3, activews.max_row + 1):

                if activews.cell(row=row,column=1) != None: 

                    for col in range(1, 13):

                        record_value = activews.cell(row=row, column=col).value

                        if record_value != None:

                            record_value = str(record_value).strip()
                            record_value = record_value.lower()

                            if record_value == "monerary":

                                record_value = "monetary"

                            if "percentage" in record_value:

                                record_value = record_value.replace("percentage", "percent")

                        record[self.field_keys[col-1]] = record_value

                    record['csrd_sheet_name'] = ws

                    self.create_record(record)

        return {
            'type': 'ir.actions.client',
            'tag': 'reload',
        }

    
    def create_record(self,record):

        document_id = self.env["document.csrd"].search([("csrd_id", "=", record["csrd_id"])])

        if not document_id:

            self.env["document.csrd"].create(record)

        

